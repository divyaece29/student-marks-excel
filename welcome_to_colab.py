
!pip install openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill

wb = Workbook()
ws1 = wb.active
ws1.title = "All Students"
ws2=wb.create_sheet(title = "Toppers")

headers = ["Student Name","Maths","Science","Total","Average","Grade"]

for ws in [ws1,ws2]:
  for col,header in enumerate(headers,start=1):
    cell = ws.cell(row=1,column=col,value=header)
    cell.font = Font(bold=True,color="FFFFFF")
    cell.fill = PatternFill(fill_type="solid",fgColor="4472C4")

n = int(input("How many students?"))
topper_row = 2
for i in range(n):
  print(f"\nStudent {i+1}:")
  name = input("ENter name:")
  maths = int(input("Enter Maths marks:"))
  science = int(input("Enter Science marks:"))

  total = maths + science
  average = total / 2

  if average >= 90:
    grade = "A+"
  elif average >= 80:
    grade = "A"
  elif average >= 70:
    grade = "B"
  elif average >= 60:
    grade = "C"
  else:
    grade = "F"

  data = [name,maths,science,total,average,grade]
  for col,value in enumerate(data,start=1):
    ws1.cell(row=i+2,column=col,value=value)

  if average >= 80:
    for col,value in enumerate(data,start=1):
      ws2.cell(row=topper_row,column=col,value=value)
    topper_row += 1

wb.save("student_marks.xlsx")
print("\nDone!File saved.")

from google.colab import files
files.download("student_marks.xlsx")
